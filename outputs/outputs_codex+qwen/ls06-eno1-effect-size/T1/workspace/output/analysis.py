"""Compute ENO1 tumor-vs-normal effect size from the proteomics workbook.

Source: inputs/Proteomic_data .xlsx, sheet "Tumor vs Normal".
Fold change (FC) = Tumor / Normal; log2FC = log2(FC).
The unrelated MeRIP_RNA_result.xlsx workbook is intentionally not used.
"""
import json
import math
from pathlib import Path

from openpyxl import load_workbook

WORKSPACE = Path(__file__).resolve().parent.parent
SOURCE_FILE = "Proteomic_data .xlsx"
SOURCE_SHEET = "Tumor vs Normal"
GENE = "ENO1"


def main() -> None:
    wb = load_workbook(WORKSPACE / "inputs" / SOURCE_FILE, read_only=True, data_only=True)
    ws = wb[SOURCE_SHEET]
    rows = ws.iter_rows(values_only=True)
    header = list(next(rows))
    idx = {name: header.index(name) for name in ("gene", "Normal", "Tumor")}

    record = None
    for row in rows:
        if row[idx["gene"]] == GENE:
            record = row
            break
    if record is None:
        raise SystemExit(f"{GENE} not found in sheet {SOURCE_SHEET!r}")

    normal = float(record[idx["Normal"]])
    tumor = float(record[idx["Tumor"]])
    if normal <= 0 or tumor <= 0:
        raise SystemExit("Non-positive abundance value; fold change undefined")

    fold_change = tumor / normal
    log2_fold_change = math.log2(fold_change)

    result = {
        "gene": GENE,
        "tumor_value": tumor,
        "normal_value": normal,
        "fold_change": fold_change,
        "log2_fold_change": log2_fold_change,
        "source_file": SOURCE_FILE,
        "source_sheet": SOURCE_SHEET,
    }
    out_path = WORKSPACE / "output" / "eno1_effect.json"
    out_path.write_text(json.dumps(result, indent=2) + "\n", encoding="utf-8")

    direction = "up in tumor" if fold_change > 1 else ("down in tumor" if fold_change < 1 else "unchanged")
    print(json.dumps(result, indent=2))
    print(f"direction: {direction}")
    # Cross-check against workbook-reported columns (rounded in the sheet).
    fc_col = float(record[header.index("FC")])
    log2_col = float(record[header.index("log2FC")])
    print(f"workbook FC={fc_col}, log2FC={log2_col} (rounded); computed FC={fold_change:.4f}, log2FC={log2_fold_change:.4f}")


if __name__ == "__main__":
    main()
