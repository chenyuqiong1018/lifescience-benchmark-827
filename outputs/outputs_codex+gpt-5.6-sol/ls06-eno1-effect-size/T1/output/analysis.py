"""Reproducible ENO1 fold-change calculation from the designated proteomics sheet."""

from __future__ import annotations

import json
import math
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


TARGET_GENE = "ENO1"
TARGET_FILE = "Proteomic_data .xlsx"
TARGET_SHEET = "Tumor vs Normal"
REQUIRED_HEADERS = (
    "gene",
    "Normal",
    "Tumor",
    "Ratio",
    "FC",
    "log2FC",
    "p.value",
    "adj.Pval",
)
JSON_KEYS = (
    "gene",
    "tumor_value",
    "normal_value",
    "fold_change",
    "log2_fold_change",
    "source_file",
    "source_sheet",
)


def finite_number(value: Any, field: str) -> float:
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        raise ValueError(f"ENO1 {field} is not numeric: {value!r}")
    converted = float(value)
    if not math.isfinite(converted):
        raise ValueError(f"ENO1 {field} is not finite: {value!r}")
    return converted


def calculate(workbook_path: Path) -> tuple[dict[str, Any], dict[str, Any]]:
    workbook = load_workbook(workbook_path, read_only=False, data_only=False)
    try:
        if workbook.sheetnames != [TARGET_SHEET]:
            raise ValueError(
                f"Expected exactly [{TARGET_SHEET!r}], found {workbook.sheetnames!r}"
            )
        sheet = workbook[TARGET_SHEET]
        if sheet.sheet_state != "visible":
            raise ValueError(f"Target sheet state is {sheet.sheet_state!r}, not visible")

        headers = [cell.value for cell in sheet[1]]
        missing_headers = [name for name in REQUIRED_HEADERS if name not in headers]
        if missing_headers:
            raise ValueError(f"Missing required headers: {missing_headers}")
        positions = {name: headers.index(name) + 1 for name in REQUIRED_HEADERS}

        matching_rows = [
            row_number
            for row_number in range(2, sheet.max_row + 1)
            if sheet.cell(row_number, positions["gene"]).value == TARGET_GENE
        ]
        if len(matching_rows) != 1:
            raise ValueError(
                f"Expected one {TARGET_GENE} row, found {len(matching_rows)}"
            )
        row_number = matching_rows[0]

        effect_fields = ("gene", "Normal", "Tumor", "Ratio", "FC", "log2FC")
        row_values = {
            name: sheet.cell(row_number, positions[name]).value for name in effect_fields
        }
        missing_values = [name for name, value in row_values.items() if value is None]
        if missing_values:
            raise ValueError(f"Missing ENO1 values: {missing_values}")
        formula_fields = [
            name
            for name, value in row_values.items()
            if isinstance(value, str) and value.startswith("=")
        ]
        if formula_fields:
            raise ValueError(f"Unexpected formulas in ENO1 effect fields: {formula_fields}")

        normal = finite_number(row_values["Normal"], "Normal")
        tumor = finite_number(row_values["Tumor"], "Tumor")
        if normal <= 0 or tumor <= 0:
            raise ValueError("Normal and Tumor must both be positive")

        fold_change = tumor / normal
        log2_fold_change = math.log2(fold_change)
        workbook_ratio = finite_number(row_values["Ratio"], "Ratio")
        workbook_fc = finite_number(row_values["FC"], "FC")
        workbook_log2fc = finite_number(row_values["log2FC"], "log2FC")

        expected_fold_display = round(fold_change, 2)
        expected_log_display = round(log2_fold_change, 2)
        for label, observed in (("Ratio", workbook_ratio), ("FC", workbook_fc)):
            if not math.isclose(observed, expected_fold_display, abs_tol=0.005):
                raise ValueError(
                    f"Workbook {label}={observed} disagrees with calculated fold change"
                )
        if not math.isclose(workbook_log2fc, expected_log_display, abs_tol=0.005):
            raise ValueError(
                "Workbook log2FC disagrees with calculated log2 fold change"
            )

        required_column_letters = [
            sheet.cell(1, positions[name]).column_letter for name in effect_fields
        ]
        hidden_required_columns = [
            letter for letter in required_column_letters if sheet.column_dimensions[letter].hidden
        ]
        if sheet.row_dimensions[row_number].hidden or hidden_required_columns:
            raise ValueError("ENO1 row or a required effect column is hidden")

        result: dict[str, Any] = {
            "gene": TARGET_GENE,
            "tumor_value": tumor,
            "normal_value": normal,
            "fold_change": fold_change,
            "log2_fold_change": log2_fold_change,
            "source_file": TARGET_FILE,
            "source_sheet": TARGET_SHEET,
        }
        if tuple(result) != JSON_KEYS:
            raise AssertionError("JSON field order/schema changed")

        checks: dict[str, Any] = {
            "sheet_count": len(workbook.sheetnames),
            "sheet_state": sheet.sheet_state,
            "rows_including_header": sheet.max_row,
            "columns": sheet.max_column,
            "eno1_row": row_number,
            "eno1_matches": len(matching_rows),
            "missing_effect_fields": missing_values,
            "formula_effect_fields": formula_fields,
            "workbook_ratio": workbook_ratio,
            "workbook_fc": workbook_fc,
            "workbook_log2fc": workbook_log2fc,
        }
        return result, checks
    finally:
        workbook.close()


def make_report(result: dict[str, Any], checks: dict[str, Any]) -> str:
    fold = float(result["fold_change"])
    log2_fold = float(result["log2_fold_change"])
    direction = "higher in tumor" if fold > 1 else "lower in tumor"
    return f"""# ENO1 tumor-versus-normal fold change

## Answer

ENO1 is **{direction}**. The tumor-versus-normal fold change is `{fold:.15g}`, and the log2 fold change is `{log2_fold:.15g}`.

| Field | Value |
|---|---:|
| Gene | {result['gene']} |
| Tumor value | {float(result['tumor_value']):.15g} |
| Normal value | {float(result['normal_value']):.15g} |
| Fold change (`Tumor / Normal`) | {fold:.15g} |
| log2 fold change | {log2_fold:.15g} |

Source: `{result['source_file']}`, sheet `{result['source_sheet']}`.

## Validation and interpretation

- The designated workbook has {checks['sheet_count']} visible sheet, {checks['rows_including_header']} rows including the header, and {checks['columns']} columns.
- ENO1 appears exactly once (worksheet row {checks['eno1_row']}); all effect fields are present, numeric, non-formula cells, and visible.
- Direct calculations agree with the workbook's rounded `Ratio` and `FC` values (`{float(checks['workbook_fc']):.2f}`) and rounded `log2FC` (`{float(checks['workbook_log2fc']):.2f}`).
- A T1-only MarkItDown conversion of this same workbook preserved the ENO1 row and the same rounded values.
- This is a descriptive fold-change calculation. The supplied summary row is not used to invent a confidence interval or a new hypothesis test, and no claim of statistical significance is made here.
- No physical unit is assigned because the input does not specify one. External proteome annotations are not mixed into this file-derived calculation.

The unrelated RNA/m6A workbook was not opened or used.
"""


def main() -> None:
    output_dir = Path(__file__).resolve().parent
    repo_root = Path(__file__).resolve().parents[5]
    workbook_path = repo_root / "inputs" / "ls06-eno1-effect-size" / TARGET_FILE
    result, checks = calculate(workbook_path)
    (output_dir / "eno1_effect.json").write_text(
        json.dumps(result, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )
    (output_dir / "report.md").write_text(
        make_report(result, checks), encoding="utf-8"
    )


if __name__ == "__main__":
    main()
