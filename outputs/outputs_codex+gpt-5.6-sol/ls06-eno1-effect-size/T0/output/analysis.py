"""Calculate the ENO1 tumor-versus-normal effect from the proteomics workbook."""

from __future__ import annotations

import json
import math
from pathlib import Path

from openpyxl import load_workbook


GENE = "ENO1"
SOURCE_FILE = "Proteomic_data .xlsx"
SOURCE_SHEET = "Tumor vs Normal"
REQUIRED_COLUMNS = ("gene", "Normal", "Tumor", "Ratio", "FC", "log2FC")
OUTPUT_FIELDS = (
    "gene",
    "tumor_value",
    "normal_value",
    "fold_change",
    "log2_fold_change",
    "source_file",
    "source_sheet",
)


def as_number(value: object, label: str) -> float:
    """Return a finite numeric value without treating booleans as measurements."""
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        raise ValueError(f"{label} must be numeric; got {value!r}")
    number = float(value)
    if not math.isfinite(number):
        raise ValueError(f"{label} must be finite; got {value!r}")
    return number


def analyze(workbook_path: Path) -> tuple[dict[str, object], dict[str, object]]:
    """Validate the target sheet, locate ENO1 uniquely, and calculate its effect."""
    workbook = load_workbook(workbook_path, read_only=True, data_only=False)
    try:
        if workbook.sheetnames != [SOURCE_SHEET]:
            raise ValueError(
                f"Expected only sheet {SOURCE_SHEET!r}; found {workbook.sheetnames!r}"
            )
        worksheet = workbook[SOURCE_SHEET]
        headers = [cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
        missing_headers = [name for name in REQUIRED_COLUMNS if name not in headers]
        if missing_headers:
            raise ValueError(f"Missing required columns: {missing_headers}")

        column = {name: headers.index(name) for name in REQUIRED_COLUMNS}
        matches: list[tuple[int, tuple[object, ...]]] = []
        for row_number, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), 2):
            if row[column["gene"]] == GENE:
                matches.append((row_number, row))

        if len(matches) != 1:
            raise ValueError(f"Expected one ENO1 row; found {len(matches)}")

        row_number, row = matches[0]
        critical = {name: row[column[name]] for name in REQUIRED_COLUMNS}
        missing_values = [name for name, value in critical.items() if value is None]
        if missing_values:
            raise ValueError(f"ENO1 has missing required values: {missing_values}")

        normal = as_number(critical["Normal"], "ENO1 Normal")
        tumor = as_number(critical["Tumor"], "ENO1 Tumor")
        if normal <= 0 or tumor <= 0:
            raise ValueError("Normal and Tumor values must be positive for ratio/log2")

        fold_change = tumor / normal
        log2_fold_change = math.log2(fold_change)

        supplied_ratio = as_number(critical["Ratio"], "ENO1 Ratio")
        supplied_fc = as_number(critical["FC"], "ENO1 FC")
        supplied_log2fc = as_number(critical["log2FC"], "ENO1 log2FC")
        if not math.isclose(supplied_ratio, round(fold_change, 2), abs_tol=0.005):
            raise ValueError("Calculated fold change disagrees with workbook Ratio")
        if not math.isclose(supplied_fc, round(fold_change, 2), abs_tol=0.005):
            raise ValueError("Calculated fold change disagrees with workbook FC")
        if not math.isclose(supplied_log2fc, round(log2_fold_change, 2), abs_tol=0.005):
            raise ValueError("Calculated log2 fold change disagrees with workbook log2FC")

        result: dict[str, object] = {
            "gene": GENE,
            "tumor_value": tumor,
            "normal_value": normal,
            "fold_change": fold_change,
            "log2_fold_change": log2_fold_change,
            "source_file": SOURCE_FILE,
            "source_sheet": SOURCE_SHEET,
        }
        if tuple(result) != OUTPUT_FIELDS:
            raise AssertionError("Output schema changed unexpectedly")

        audit: dict[str, object] = {
            "sheet_count": len(workbook.sheetnames),
            "sheet_state": worksheet.sheet_state,
            "row_count_including_header": worksheet.max_row,
            "column_count": worksheet.max_column,
            "eno1_row": row_number,
            "eno1_match_count": len(matches),
            "missing_required_values": missing_values,
            "workbook_ratio": supplied_ratio,
            "workbook_fc": supplied_fc,
            "workbook_log2fc": supplied_log2fc,
        }
        return result, audit
    finally:
        workbook.close()


def render_report(result: dict[str, object], audit: dict[str, object]) -> str:
    direction = "higher in tumor" if result["fold_change"] > 1 else "lower in tumor"
    return f"""# ENO1 tumor-versus-normal effect size

## Result

ENO1 is **{direction}**. Using the supplied abundance values, the tumor-versus-normal fold change is `{result['fold_change']:.15g}` and the log2 fold change is `{result['log2_fold_change']:.15g}`.

- Tumor value: `{result['tumor_value']:.15g}`
- Normal value: `{result['normal_value']:.15g}`
- Calculation: `Tumor / Normal`
- Source file: `{result['source_file']}`
- Source sheet: `{result['source_sheet']}`

## Focused data checks

- The target workbook contains {audit['sheet_count']} visible analysis sheet with {audit['row_count_including_header']} rows including the header and {audit['column_count']} columns.
- ENO1 occurs exactly once, at worksheet row {audit['eno1_row']}.
- The ENO1 `gene`, `Normal`, `Tumor`, `Ratio`, `FC`, and `log2FC` cells are all present.
- The direct calculation agrees with the workbook's rounded `Ratio`/`FC` (`{audit['workbook_fc']:.2f}`) and `log2FC` (`{audit['workbook_log2fc']:.2f}`) values.

The workbook does not specify a physical unit for the abundance values, so none is invented. No inferential significance claim is made from this single requested effect-size calculation. The unrelated RNA/m6A workbook was not opened or used.
"""


def main() -> None:
    repo_root = Path(__file__).resolve().parents[5]
    workbook_path = repo_root / "inputs" / "ls06-eno1-effect-size" / SOURCE_FILE
    output_dir = Path(__file__).resolve().parent
    result, audit = analyze(workbook_path)

    (output_dir / "eno1_effect.json").write_text(
        json.dumps(result, indent=2, ensure_ascii=False) + "\n", encoding="utf-8"
    )
    (output_dir / "report.md").write_text(
        render_report(result, audit), encoding="utf-8"
    )


if __name__ == "__main__":
    main()
