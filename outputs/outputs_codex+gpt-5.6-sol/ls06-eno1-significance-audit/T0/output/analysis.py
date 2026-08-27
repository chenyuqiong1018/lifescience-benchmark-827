"""Threshold-calibrated audit of ENO1's adjusted proteomics p-value."""

from __future__ import annotations

import json
import math
from pathlib import Path

from openpyxl import load_workbook


GENE = "ENO1"
SOURCE_FILE = "Proteomic_data .xlsx"
SOURCE_SHEET = "Tumor vs Normal"
FDR_THRESHOLD = 0.05
OUTPUT_FIELDS = (
    "gene",
    "adjusted_p_value",
    "fdr_threshold",
    "significant",
    "source_file",
    "source_sheet",
)


def valid_probability(value: object, label: str) -> float:
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        raise ValueError(f"{label} must be numeric; got {value!r}")
    number = float(value)
    if not math.isfinite(number) or not 0 <= number <= 1:
        raise ValueError(f"{label} must be finite and within [0,1]; got {value!r}")
    return number


def analyze(workbook_path: Path) -> tuple[dict[str, object], dict[str, object]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=False)
    try:
        if workbook.sheetnames != [SOURCE_SHEET]:
            raise ValueError(
                f"Expected only sheet {SOURCE_SHEET!r}; found {workbook.sheetnames!r}"
            )
        sheet = workbook[SOURCE_SHEET]
        headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        required = ("gene", "p.value", "adj.Pval")
        missing_headers = [name for name in required if name not in headers]
        if missing_headers:
            raise ValueError(f"Missing required columns: {missing_headers}")
        positions = {name: headers.index(name) for name in required}
        if positions["p.value"] == positions["adj.Pval"]:
            raise ValueError("Raw and adjusted p-values are not distinct columns")

        matches = []
        for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):
            if row[positions["gene"]] == GENE:
                matches.append((row_number, row))
        if len(matches) != 1:
            raise ValueError(f"Expected one ENO1 row; found {len(matches)}")

        row_number, row = matches[0]
        raw_p = valid_probability(row[positions["p.value"]], "ENO1 raw p.value")
        adjusted_p = valid_probability(
            row[positions["adj.Pval"]], "ENO1 adjusted p-value"
        )
        result: dict[str, object] = {
            "gene": GENE,
            "adjusted_p_value": adjusted_p,
            "fdr_threshold": FDR_THRESHOLD,
            "significant": adjusted_p <= FDR_THRESHOLD,
            "source_file": SOURCE_FILE,
            "source_sheet": SOURCE_SHEET,
        }
        if tuple(result) != OUTPUT_FIELDS:
            raise AssertionError("Output schema changed unexpectedly")
        audit: dict[str, object] = {
            "raw_p_value": raw_p,
            "eno1_row": row_number,
            "eno1_matches": len(matches),
            "rows_including_header": sheet.max_row,
            "columns": sheet.max_column,
        }
        return result, audit
    finally:
        workbook.close()


def render_report(result: dict[str, object], audit: dict[str, object]) -> str:
    outcome = "passes" if result["significant"] else "does not pass"
    return f"""# ENO1 FDR significance audit

## Threshold-calibrated result

ENO1's adjusted p-value is `{result['adjusted_p_value']}`. It **{outcome}** the prespecified FDR `{result['fdr_threshold']}` threshold, so `significant` is `{str(result['significant']).lower()}`.

- Gene: `{result['gene']}`
- Adjusted column used: `adj.Pval`
- Adjusted p-value: `{result['adjusted_p_value']}`
- Decision rule: `adjusted_p_value <= 0.05`
- Source: `{result['source_file']}`, sheet `{result['source_sheet']}`

## Statistical interpretation

The raw `p.value` in the ENO1 row is `{audit['raw_p_value']}`, which is below 0.05, but it is not the requested multiplicity-adjusted value. The adjusted value `{result['adjusted_p_value']}` is above 0.05 and therefore does not meet the requested FDR criterion. This threshold result is not proof of no biological effect; it means only that ENO1 is not significant under the specified adjusted-p-value rule.

ENO1 occurs exactly once (worksheet row {audit['eno1_row']}) in the target sheet. No new hypothesis test was invented, and no worksheet flag was substituted for the explicit FDR comparison. The unrelated RNA/m6A workbook was not opened or used.
"""


def main() -> None:
    output_dir = Path(__file__).resolve().parent
    repo_root = Path(__file__).resolve().parents[5]
    workbook_path = repo_root / "inputs" / "ls06-eno1-significance-audit" / SOURCE_FILE
    result, audit = analyze(workbook_path)
    (output_dir / "eno1_significance.json").write_text(
        json.dumps(result, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )
    (output_dir / "report.md").write_text(
        render_report(result, audit), encoding="utf-8"
    )


if __name__ == "__main__":
    main()
