"""Reproduce ENO1's adjusted-p-value decision at an FDR threshold of 0.05."""

from __future__ import annotations

import json
import math
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


TARGET_GENE = "ENO1"
TARGET_FILE = "Proteomic_data .xlsx"
TARGET_SHEET = "Tumor vs Normal"
FDR_THRESHOLD = 0.05
REQUIRED_HEADERS = ("protein", "gene", "p.value", "adj.Pval", "is_sig")
OUTPUT_KEYS = (
    "gene",
    "adjusted_p_value",
    "fdr_threshold",
    "significant",
    "source_file",
    "source_sheet",
)


def probability(value: Any, label: str) -> float:
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        raise ValueError(f"{label} is not numeric: {value!r}")
    number = float(value)
    if not math.isfinite(number) or not 0 <= number <= 1:
        raise ValueError(f"{label} is not a finite probability in [0,1]: {value!r}")
    return number


def inspect_target(workbook_path: Path) -> tuple[dict[str, Any], dict[str, Any]]:
    workbook = load_workbook(workbook_path, read_only=False, data_only=False)
    try:
        if workbook.sheetnames != [TARGET_SHEET]:
            raise ValueError(
                f"Expected exactly [{TARGET_SHEET!r}], found {workbook.sheetnames!r}"
            )
        sheet = workbook[TARGET_SHEET]
        if sheet.sheet_state != "visible":
            raise ValueError(f"Target sheet is not visible: {sheet.sheet_state!r}")
        headers = [cell.value for cell in sheet[1]]
        missing_headers = [name for name in REQUIRED_HEADERS if name not in headers]
        if missing_headers:
            raise ValueError(f"Missing required headers: {missing_headers}")
        columns = {name: headers.index(name) + 1 for name in REQUIRED_HEADERS}
        if columns["p.value"] == columns["adj.Pval"]:
            raise ValueError("Raw and adjusted p-values resolve to the same column")

        matching_rows = [
            row_number
            for row_number in range(2, sheet.max_row + 1)
            if sheet.cell(row_number, columns["gene"]).value == TARGET_GENE
        ]
        if len(matching_rows) != 1:
            raise ValueError(
                f"Expected one {TARGET_GENE} row; found {len(matching_rows)}"
            )
        row_number = matching_rows[0]

        raw_cell = sheet.cell(row_number, columns["p.value"])
        adjusted_cell = sheet.cell(row_number, columns["adj.Pval"])
        if any(
            isinstance(cell.value, str) and cell.value.startswith("=")
            for cell in (raw_cell, adjusted_cell)
        ):
            raise ValueError("ENO1 raw/adjusted p-value cell unexpectedly contains a formula")
        raw_p = probability(raw_cell.value, "ENO1 raw p.value")
        adjusted_p = probability(adjusted_cell.value, "ENO1 adj.Pval")
        worksheet_flag = sheet.cell(row_number, columns["is_sig"]).value

        result: dict[str, Any] = {
            "gene": TARGET_GENE,
            "adjusted_p_value": adjusted_p,
            "fdr_threshold": FDR_THRESHOLD,
            "significant": adjusted_p <= FDR_THRESHOLD,
            "source_file": TARGET_FILE,
            "source_sheet": TARGET_SHEET,
        }
        if tuple(result) != OUTPUT_KEYS:
            raise AssertionError("Output JSON schema changed")

        required_letters = [
            sheet.cell(1, columns[name]).column_letter for name in REQUIRED_HEADERS
        ]
        hidden_required_columns = [
            letter for letter in required_letters if sheet.column_dimensions[letter].hidden
        ]
        if sheet.row_dimensions[row_number].hidden or hidden_required_columns:
            raise ValueError("ENO1 row or a required audit column is hidden")

        audit: dict[str, Any] = {
            "raw_p_value": raw_p,
            "worksheet_flag": worksheet_flag,
            "eno1_row": row_number,
            "eno1_match_count": len(matching_rows),
            "rows_including_header": sheet.max_row,
            "columns": sheet.max_column,
            "sheet_state": sheet.sheet_state,
        }
        return result, audit
    finally:
        workbook.close()


def make_report(result: dict[str, Any], audit: dict[str, Any]) -> str:
    decision = "significant" if result["significant"] else "not significant"
    return f"""# ENO1 adjusted-p-value significance audit

## Result

ENO1's adjusted p-value is `{result['adjusted_p_value']}`. At FDR `{result['fdr_threshold']}`, ENO1 is **{decision}** because the decision expression `{result['adjusted_p_value']} <= {result['fdr_threshold']}` evaluates to `{str(result['significant']).lower()}`.

| Field | Value |
|---|---|
| Gene | `{result['gene']}` |
| Adjusted p-value (`adj.Pval`) | `{result['adjusted_p_value']}` |
| FDR threshold | `{result['fdr_threshold']}` |
| Significant | `{str(result['significant']).lower()}` |
| Source file | `{result['source_file']}` |
| Source sheet | `{result['source_sheet']}` |

## Provenance and audit checks

- ENO1 appears exactly once at worksheet row {audit['eno1_row']} in the single visible target sheet.
- The raw `p.value` (`{audit['raw_p_value']}`) and adjusted `adj.Pval` (`{result['adjusted_p_value']}`) are distinct columns and values. The raw value is not relabeled as adjusted.
- The worksheet's existing `is_sig` value is `{audit['worksheet_flag']}`, but it is not used for this answer because it conflicts with the requested direct rule `adj.Pval <= 0.05`.
- A T1-only MarkItDown conversion of the same target workbook preserved the ENO1 row and both p-value columns.
- The controlled code-execution handoff used the same threshold expression; the saved local script performs the authoritative executable calculation.

This result means ENO1 does not meet the specified FDR criterion; it is not proof that ENO1 has no biological effect. No additional hypothesis test was invented. External proteome and biomarker databases were not mixed into this workbook-derived significance decision. The unrelated RNA/m6A workbook was not opened or used.
"""


def main() -> None:
    output_dir = Path(__file__).resolve().parent
    repo_root = Path(__file__).resolve().parents[5]
    workbook_path = repo_root / "inputs" / "ls06-eno1-significance-audit" / TARGET_FILE
    result, audit = inspect_target(workbook_path)
    (output_dir / "eno1_significance.json").write_text(
        json.dumps(result, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )
    (output_dir / "report.md").write_text(
        make_report(result, audit), encoding="utf-8"
    )


if __name__ == "__main__":
    main()
